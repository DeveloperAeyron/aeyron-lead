"""Batch email generation from spreadsheet rows (research + Ollama)."""
from __future__ import annotations

import csv
import io
import threading
import time
from io import BytesIO
from typing import Any, Callable, Mapping

from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook

from email_generation import (
    EmailGenerationInput,
    _research_namespace,
    company_cache_key,
    generate_with_trimmed_context,
    make_prompt_row,
    prepend_prospect_website_block,
    split_full_name,
)
from generate_news_emails_ollama import (
    _company_report_context,
    _research_company,
    _trim_news_context,
)

_EMAIL_KEYS = ("email", "e-mail", "e mail", "email address", "mail")
_NAME_KEYS = ("name", "full name", "contact name", "recipient name", "contact")
_FIRST_KEYS = ("first name", "firstname", "given name", "first")
_LAST_KEYS = ("last name", "lastname", "surname", "family name", "last")
_COMPANY_KEYS = (
    "company name",
    "company",
    "organisation",
    "organization",
    "org",
    "business name",
    "employer",
)
_POSITION_KEYS = ("position", "title", "job title", "role")
_INDUSTRY_KEYS = ("industry", "sector", "vertical")
_WEBSITE_KEYS = ("website", "web site", "company website", "company url", "domain", "site")


def _norm_header(h: str) -> str:
    return " ".join((h or "").strip().lower().split())


def _pick_field(row: Mapping[str, Any], candidates: tuple[str, ...]) -> str:
    """Match spreadsheet column to canonical field (case-insensitive, flexible naming)."""
    lower_map = {_norm_header(str(k)): v for k, v in row.items()}
    for c in candidates:
        if c in lower_map:
            v = lower_map[c]
            if v is not None and str(v).strip():
                return str(v).strip()
    # Substring fallbacks
    cand_set = set(candidates)
    for k, v in lower_map.items():
        if v is None or not str(v).strip():
            continue
        for c in cand_set:
            if c in k or (len(c) >= 5 and k in c):
                return str(v).strip()
    return ""


def extract_row_for_prompt(
    row: Mapping[str, Any], base: EmailGenerationInput
) -> tuple[str, str, str, str, str, str]:
    """email, first, last, company, position, industry (per-row titles from CSV/Excel when present)."""
    email = _pick_field(row, _EMAIL_KEYS)
    company = _pick_field(row, _COMPANY_KEYS)
    fn = _pick_field(row, _FIRST_KEYS)
    ln = _pick_field(row, _LAST_KEYS)
    if not fn and not ln:
        full = _pick_field(row, _NAME_KEYS)
        fn, ln = split_full_name(full)
    position = _pick_field(row, _POSITION_KEYS) or (base.position or "")
    industry = _pick_field(row, _INDUSTRY_KEYS) or (base.industry or "")
    return email, fn, ln, company, position, industry


def load_rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Read the first worksheet; first row = headers. All columns preserved."""
    bio = BytesIO(content)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        headers = [str(c).strip() if c is not None else "" for c in header_row]
        out: list[dict[str, Any]] = []
        for row in rows_iter:
            if row is None:
                continue
            if all((c is None or str(c).strip() == "") for c in row):
                continue
            record: dict[str, Any] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                val = row[i] if i < len(row) else None
                record[h] = "" if val is None else str(val).strip()
            out.append(record)
        return out
    finally:
        wb.close()


def load_rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    """Parse CSV (comma, semicolon, or tab). UTF-8 with BOM and Latin-1 fallback; supports quoted multiline fields."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    sample = text[: min(16_384, len(text))]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    f = io.StringIO(text)
    reader = csv.DictReader(f, dialect=dialect)
    out: list[dict[str, Any]] = []
    for r in reader:
        if not r:
            continue
        if all((v is None or str(v).strip() == "") for v in r.values()):
            continue
        cleaned: dict[str, Any] = {}
        for k, v in r.items():
            if k is None:
                continue
            key = str(k).strip()
            if not key:
                continue
            cleaned[key] = "" if v is None else str(v).strip()
        if cleaned:
            out.append(cleaned)
    return out


def process_batch(
    rows: list[dict[str, Any]],
    base: EmailGenerationInput,
    *,
    max_rows: int = 50,
    sleep_ms: int = 0,
    concurrency: int = 1,
    on_progress: Callable[[int, int, str], None] | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    For each row: company research (cached per company), personalised email, news summary.
    Returns (result dicts for CSV, list of fieldnames order).
    """
    if not rows:
        raise ValueError("No data rows in spreadsheet.")

    capped = rows[: max(1, max_rows)]
    total = len(capped)
    research_cache: dict[str, list[dict[str, str]]] = {}
    research_locks: dict[str, threading.Lock] = {}
    research_locks_guard = threading.Lock()
    outputs: list[dict[str, Any]] = []

    # Stable column order: original keys from first row, then new columns
    base_keys = list(capped[0].keys())
    extra_cols = ["subject", "body", "news based summary", "generation_error"]
    fieldnames = base_keys + [c for c in extra_cols if c not in base_keys]

    def _company_lock(cache_key: str) -> threading.Lock:
        # Lazily create a per-company lock, guarded by a global lock.
        with research_locks_guard:
            lk = research_locks.get(cache_key)
            if lk is None:
                lk = threading.Lock()
                research_locks[cache_key] = lk
            return lk

    def _process_one(i: int, raw: dict[str, Any]) -> tuple[int, dict[str, Any]]:
        line_out = {k: raw.get(k, "") for k in base_keys}
        for c in extra_cols:
            line_out.setdefault(c, "")

        email, first, last, company, position, industry = extract_row_for_prompt(raw, base)

        if not company:
            line_out["generation_error"] = "Missing company name (use a column named e.g. company name or company)."
            return i, line_out

        row_dict = make_prompt_row(
            email=email,
            first_name=first,
            last_name=last,
            company_name=company,
            position=position,
            industry=industry,
        )

        raw_news: list[dict[str, str]] = []
        research_used = False
        try:
            ck = company_cache_key(company)
            if ck:
                # Ensure only one thread performs research per company per run.
                with _company_lock(ck):
                    if ck in research_cache:
                        raw_news = list(research_cache[ck])
                    else:
                        if base.do_research:
                            report = _research_company(company, _research_namespace(base))
                            raw_news = list(_company_report_context(report))
                            research_cache[ck] = list(raw_news)
                            research_used = True
                        else:
                            raw_news = []
            else:
                if base.do_research:
                    report = _research_company(company, _research_namespace(base))
                    raw_news = list(_company_report_context(report))
                    research_used = True
                else:
                    raw_news = []

            row_website = _pick_field(raw, _WEBSITE_KEYS)
            prepend_prospect_website_block(raw_news, row_website)

            trimmed = _trim_news_context(
                raw_news,
                max_items=max(1, base.context_items),
                snippet_chars=base.context_snippet_chars,
            )
            result = generate_with_trimmed_context(
                row_dict, trimmed, base,
                research_used=research_used,
                website_url_override=row_website,
            )
            line_out["subject"] = result.get("subject", "")
            line_out["body"] = result.get("email_body", "")
            line_out["news based summary"] = result.get("news_based_summary", "")
            line_out["generation_error"] = ""
        except Exception as exc:
            line_out["generation_error"] = str(exc)

        return i, line_out

    conc = max(1, int(concurrency))
    if conc == 1:
        for i, raw in enumerate(capped):
            _, line_out = _process_one(i, raw)
            outputs.append(line_out)
            if on_progress is not None:
                company = _pick_field(raw, _COMPANY_KEYS) or ""
                try:
                    on_progress(i + 1, total, company)
                except Exception:
                    pass
            if sleep_ms > 0 and i + 1 < len(capped):
                time.sleep(sleep_ms / 1000.0)
        return outputs, fieldnames

    # Parallel execution (bounded by conc). We preserve output row order.
    ordered: list[dict[str, Any] | None] = [None] * len(capped)
    with ThreadPoolExecutor(max_workers=conc) as ex:
        futs = []
        for i, raw in enumerate(capped):
            futs.append(ex.submit(_process_one, i, raw))
            if sleep_ms > 0 and i + 1 < len(capped):
                # Rate-limit submission to avoid spiky upstream requests.
                time.sleep(sleep_ms / 1000.0)

        completed = 0
        for fut in as_completed(futs):
            i, line_out = fut.result()
            ordered[i] = line_out
            completed += 1
            if on_progress is not None:
                company = _pick_field(capped[i], _COMPANY_KEYS) or ""
                try:
                    on_progress(completed, total, company)
                except Exception:
                    pass

    outputs = [r for r in ordered if r is not None]

    return outputs, fieldnames


def rows_to_csv_bytes(rows: list[dict[str, Any]], fieldnames: list[str]) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore", restval="")
    w.writeheader()
    w.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")
