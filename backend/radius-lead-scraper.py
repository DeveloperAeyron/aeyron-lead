#!/usr/bin/env python3
from __future__ import annotations

import argparse
import logging
import math
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Optional
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

LOG = logging.getLogger("radius-lead-scraper")


@dataclass(frozen=True)
class Lead:
    seed_query: str
    seed_location: str
    seed_name: Optional[str] = None
    seed_lat: Optional[float] = None
    seed_lng: Optional[float] = None
    radius_km: float = 1.0
    name: Optional[str] = None
    rating: Optional[str] = None
    reviews: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    plus_code: Optional[str] = None
    maps_url: Optional[str] = None
    place_lat: Optional[float] = None
    place_lng: Optional[float] = None
    distance_km: Optional[float] = None
    scraped_at_iso: str = field(
        default_factory=lambda: datetime.now(timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


_EMAIL_RE = re.compile(r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,}\b")


def _clean_whitespace(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    v = re.sub(r"\s+", " ", value).strip()
    v = re.sub(r"^[^A-Za-z0-9+(-]+", "", v).strip()
    return v or None


def _safe_inner_text(locator) -> Optional[str]:
    try:
        return _clean_whitespace(locator.inner_text(timeout=1500))
    except Exception:
        return None


def _extract_from_aria_label(page, label_contains: str) -> Optional[str]:
    candidates = page.locator(
        f'[role="button"][aria-label*="{label_contains}"], [aria-label*="{label_contains}"][role="link"]'
    )
    try:
        if candidates.count() < 1:
            return None
    except Exception:
        return None

    aria = candidates.first.get_attribute("aria-label") or ""
    aria = _clean_whitespace(aria)
    if not aria:
        return None
    parts = aria.split(":", 1)
    if len(parts) == 2:
        return _clean_whitespace(parts[1])
    return aria


def _extract_from_data_item_id(page, key: str) -> Optional[str]:
    loc = page.locator(f'button[data-item-id*="{key}"], a[data-item-id*="{key}"]').first
    return _clean_whitespace(_safe_inner_text(loc))


def _first_attr(page, selector: str, attr: str) -> Optional[str]:
    loc = page.locator(selector).first
    try:
        val = loc.get_attribute(attr, timeout=1500)
        return _clean_whitespace(val)
    except Exception:
        return None


def _first_attr_with_digits(page, selector: str, attr: str) -> Optional[str]:
    loc = page.locator(selector)
    try:
        n = min(loc.count(), 10)
    except Exception:
        return None
    for i in range(n):
        try:
            val = _clean_whitespace(loc.nth(i).get_attribute(attr, timeout=1500))
        except Exception:
            continue
        if val and re.search(r"\d", val):
            return val
    return None


def _extract_email_from_maps_page(page) -> Optional[str]:
    mailto = page.locator('a[href^="mailto:"]').first
    try:
        href = mailto.get_attribute("href", timeout=800)
        if href:
            m = _EMAIL_RE.search(href)
            if m:
                return _clean_whitespace(m.group(0))
    except Exception:
        pass

    try:
        body_text = page.locator("body").inner_text(timeout=1500)
        m = _EMAIL_RE.search(body_text or "")
        if m:
            return _clean_whitespace(m.group(0))
    except Exception:
        return None

    return None


def _wait_for_results(page, timeout_ms: int) -> None:
    page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    page.wait_for_timeout(250)
    try:
        page.wait_for_selector('a[href*="/maps/place/"], div[role="feed"], div[role="main"] h1', timeout=timeout_ms)
    except PlaywrightTimeoutError:
        pass


def _normalise_maps_url(url: str) -> str:
    return url.split("#", 1)[0]


def _parse_coords_from_url(url: str) -> tuple[Optional[float], Optional[float]]:
    """
    Typical Google Maps place URLs contain either:
    - /@lat,lng,zoomz
    - ...!3dLAT!4dLNG...
    """
    m = re.search(r"/@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?),", url)
    if m:
        return float(m.group(1)), float(m.group(2))
    m = re.search(r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)", url)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def _haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    r = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _collect_place_urls(page, limit: int) -> list[str]:
    selectors = [
        'a.hfpxzc[href*="/maps/place/"]',
        'a[href*="/maps/place/"][aria-label]',
        'a[href*="/maps/place/"]',
    ]

    seen: set[str] = set()
    urls: list[str] = []

    def grab_once() -> None:
        for sel in selectors:
            loc = page.locator(sel)
            try:
                n = min(loc.count(), 80)
            except Exception:
                continue
            for i in range(n):
                href = loc.nth(i).get_attribute("href")
                if not href:
                    continue
                if href.startswith("/"):
                    href = "https://www.google.com" + href
                href = _normalise_maps_url(href)
                if "/maps/place/" not in href:
                    continue
                if href in seen:
                    continue
                seen.add(href)
                urls.append(href)
                if len(urls) >= limit:
                    return

    grab_once()
    if len(urls) >= limit:
        return urls[:limit]

    results_feed = page.locator('div[role="feed"]').first
    no_new = 0
    max_no_new = 8

    for _ in range(250):  # safety cap
        if len(urls) >= limit or no_new >= max_no_new:
            break
        before = len(urls)
        try:
            if results_feed.is_visible(timeout=500):
                results_feed.evaluate("(el) => { el.scrollBy(0, el.scrollHeight); }")
            else:
                page.mouse.wheel(0, 1400)
        except Exception:
            page.mouse.wheel(0, 1400)
        page.wait_for_timeout(650)
        grab_once()
        no_new = no_new + 1 if len(urls) == before else 0

    return urls[:limit]


def _scrape_place_fields(page, *, seed_query: str, seed_location: str, radius_km: float, seed_name: Optional[str], seed_lat: float, seed_lng: float) -> Lead:
    name = None
    for loc in [page.locator("h1.DUwDvf").first, page.locator('div[role="main"] h1').first, page.locator("h1").first]:
        candidate = _safe_inner_text(loc)
        if candidate and candidate.lower() != "results":
            name = candidate
            break
    if not name:
        try:
            title = _clean_whitespace(page.title())
            if title and " - Google Maps" in title:
                name = _clean_whitespace(title.replace(" - Google Maps", ""))
        except Exception:
            pass

    rating = (
        _safe_inner_text(page.locator('div.F7nice span[aria-hidden="true"]').first)
        or _first_attr(page, 'span[aria-label*="stars"]', "aria-label")
        or _safe_inner_text(page.locator('span[aria-label*="stars"]').first)
    )
    if rating:
        m = re.search(r"([0-9.]+)\s*stars", rating, re.I)
        rating = m.group(1) if m else _clean_whitespace(rating)

    reviews = (
        _first_attr_with_digits(page, 'button[aria-label*="reviews"]', "aria-label")
        or _extract_from_aria_label(page, "Reviews")
        or _extract_from_aria_label(page, "reviews")
    )
    if reviews:
        m = re.search(r"(\d[\d,]*)", reviews)
        reviews = m.group(1) if m else reviews

    address = _extract_from_aria_label(page, "Address") or _extract_from_data_item_id(page, "address")
    phone = _extract_from_aria_label(page, "Phone") or _extract_from_data_item_id(page, "phone")
    website = _extract_from_aria_label(page, "Website") or _extract_from_data_item_id(page, "authority")
    plus_code = _extract_from_aria_label(page, "Plus code") or _extract_from_data_item_id(page, "oloc")
    email = _extract_email_from_maps_page(page)

    place_lat, place_lng = _parse_coords_from_url(page.url)
    distance_km: Optional[float] = None
    if place_lat is not None and place_lng is not None:
        distance_km = _haversine_km(seed_lat, seed_lng, place_lat, place_lng)

    return Lead(
        seed_query=seed_query,
        seed_location=seed_location,
        seed_name=seed_name,
        seed_lat=seed_lat,
        seed_lng=seed_lng,
        radius_km=radius_km,
        name=name,
        rating=rating,
        reviews=reviews,
        address=address,
        phone=phone,
        website=website,
        email=email,
        plus_code=plus_code,
        maps_url=page.url,
        place_lat=place_lat,
        place_lng=place_lng,
        distance_km=distance_km,
    )


def write_xlsx(leads: Iterable[Lead], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "leads"

    leads_list = list(leads)
    if not leads_list:
        leads_list = [Lead(seed_query="", seed_location="")]

    headers = list(asdict(leads_list[0]).keys())
    ws.append(headers)
    for lead in leads_list:
        row = asdict(lead)
        ws.append([row.get(h) for h in headers])

    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx).value
            max_len = max(max_len, len(str(v or "")))
        ws.column_dimensions[col].width = min(max(12, max_len + 2), 70)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def scrape_with_radius(
    *,
    seed_query: str,
    seed_location: str,
    nearby_query: str,
    radius_km: float,
    limit: int,
    headless: bool,
    slow_mo_ms: int,
    timeout_ms: int,
    zoom: int,
) -> list[Lead]:
    started = datetime.now(timezone.utc)
    LOG.info(
        'Seed search: query="%s" location="%s" radius=%.2fkm limit=%s',
        seed_query,
        seed_location,
        radius_km,
        limit,
    )

    seed_search_text = f"{seed_query} in {seed_location}".strip()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms)
        context = browser.new_context(locale="en-GB", viewport={"width": 1400, "height": 900})
        page = context.new_page()
        page.set_default_timeout(timeout_ms)

        seed_url = f"https://www.google.com/maps/search/{quote_plus(seed_search_text)}"
        page.goto(seed_url, wait_until="domcontentloaded", timeout=timeout_ms)
        _wait_for_results(page, timeout_ms)

        seed_place_urls = _collect_place_urls(page, limit=1)
        if not seed_place_urls:
            raise RuntimeError("No seed place URL found from the initial search.")

        page.goto(seed_place_urls[0], wait_until="domcontentloaded", timeout=timeout_ms)
        page.wait_for_timeout(800)
        seed_name = _safe_inner_text(page.locator("h1.DUwDvf").first) or _safe_inner_text(page.locator("h1").first)
        seed_lat, seed_lng = _parse_coords_from_url(page.url)
        if seed_lat is None or seed_lng is None:
            raise RuntimeError("Could not extract seed coordinates from the seed place URL.")

        LOG.info('Seed: "%s" at (%.6f, %.6f)', seed_name or "(unknown)", seed_lat, seed_lng)

        # Nearby search, centred on the seed coordinate. We later compute distance and filter to radius_km.
        nearby_search_text = nearby_query.strip() or seed_query.strip()
        nearby_url = (
            f"https://www.google.com/maps/search/{quote_plus(nearby_search_text)}/@{seed_lat},{seed_lng},{zoom}z"
        )
        page.goto(nearby_url, wait_until="domcontentloaded", timeout=timeout_ms)
        _wait_for_results(page, timeout_ms)

        # Collect more than limit initially so we can filter by radius.
        candidate_urls = _collect_place_urls(page, limit=max(limit * 4, limit))
        LOG.info("Collected %s nearby candidate URL(s)", len(candidate_urls))

        # Pre-filter by distance using coordinates embedded in the candidate URL.
        # This avoids opening lots of pages (more stable + faster).
        within: list[tuple[str, float]] = []
        unknown_coord: list[str] = []
        seen: set[str] = set()
        for url in candidate_urls:
            if url in seen:
                continue
            seen.add(url)
            lat, lng = _parse_coords_from_url(url)
            if lat is None or lng is None:
                unknown_coord.append(url)
                continue
            d = _haversine_km(seed_lat, seed_lng, lat, lng)
            if d <= radius_km:
                within.append((url, d))

        within.sort(key=lambda t: t[1])
        LOG.info("Candidates within %.2f km by URL coords: %s", radius_km, len(within))

        leads: list[Lead] = []

        def scrape_url(url: str) -> Optional[Lead]:
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
                page.wait_for_timeout(500)
                return _scrape_place_fields(
                    page,
                    seed_query=seed_query,
                    seed_location=seed_location,
                    radius_km=radius_km,
                    seed_name=seed_name,
                    seed_lat=seed_lat,
                    seed_lng=seed_lng,
                )
            except Exception:
                LOG.exception("Failed to scrape candidate")
                return None

        # First, scrape the ones we are confident are within radius.
        for idx, (url, dist) in enumerate(within, start=1):
            if len(leads) >= limit:
                break
            elapsed_s = (datetime.now(timezone.utc) - started).total_seconds()
            LOG.info("Scraping within-radius %s/%s (kept %s, elapsed %.1fs)", idx, len(within), len(leads), elapsed_s)
            lead = scrape_url(url)
            if not lead:
                continue
            if lead.distance_km is not None and lead.distance_km <= radius_km:
                leads.append(lead)

        # If we still haven't reached `limit`, optionally try a few without embedded coords.
        if len(leads) < limit and unknown_coord:
            LOG.info("Trying %s URL(s) with unknown coords", len(unknown_coord))
            for idx, url in enumerate(unknown_coord, start=1):
                if len(leads) >= limit:
                    break
                elapsed_s = (datetime.now(timezone.utc) - started).total_seconds()
                LOG.info("Scraping unknown-coord %s/%s (kept %s, elapsed %.1fs)", idx, len(unknown_coord), len(leads), elapsed_s)
                lead = scrape_url(url)
                if not lead:
                    continue
                if lead.distance_km is not None and lead.distance_km <= radius_km:
                    leads.append(lead)

        context.close()
        browser.close()

        LOG.info("Kept %s lead(s) within %.2f km", len(leads), radius_km)
        return leads


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Seed a place, then scrape leads within a radius (km).")
    parser.add_argument("--seed-query", default="car wash", help='Seed search term, e.g. "car wash"')
    parser.add_argument("--seed-location", default="Delaware", help='Seed search location, e.g. "Delaware"')
    parser.add_argument("--nearby-query", default=None, help='Nearby search term (defaults to seed-query)')
    parser.add_argument("--radius-km", type=float, default=1.0, help="Radius in kilometres (default 1.0)")
    parser.add_argument("--limit", type=int, default=10, help="Number of leads to keep within radius (default 10)")
    parser.add_argument("--zoom", type=int, default=15, help="Map zoom used for the nearby search (default 15)")
    parser.add_argument("--headless", action="store_true", help="Run headless (default is headed)")
    parser.add_argument("--slowmo-ms", type=int, default=50, help="Slow motion delay per action (ms)")
    parser.add_argument("--timeout-ms", type=int, default=90_000, help="Playwright timeout (ms)")
    parser.add_argument("--out", default="radius-leads.xlsx", help="Output .xlsx path")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    nearby_query = args.nearby_query if args.nearby_query is not None else args.seed_query

    leads = scrape_with_radius(
        seed_query=args.seed_query,
        seed_location=args.seed_location,
        nearby_query=nearby_query,
        radius_km=args.radius_km,
        limit=args.limit,
        headless=args.headless,
        slow_mo_ms=args.slowmo_ms,
        timeout_ms=args.timeout_ms,
        zoom=args.zoom,
    )

    out_path = Path(args.out).expanduser().resolve()
    write_xlsx(leads, out_path)
    print(f"Wrote {len(leads)} lead(s) to: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

