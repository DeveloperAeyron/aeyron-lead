#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen


EMAIL_RE = re.compile(r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,}\b")


def _require(module_name: str) -> Any:
    try:
        return __import__(module_name)
    except ModuleNotFoundError as exc:
        raise SystemExit(
            f'Missing dependency "{module_name}".\n'
            f"- If you use a venv, activate it first.\n"
            f"- Otherwise install deps: pip install -r requirements.txt\n"
        ) from exc


def _normalise_url(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    u = str(url).strip()
    if not u:
        return None
    if u.startswith(("http://", "https://")):
        return u
    # Handle "example.com" style
    if re.match(r"^[a-zA-Z0-9.-]+\.[A-Za-z]{2,}(/|$)", u):
        return "https://" + u
    return None


def _strip_html_for_llm(html: str, *, max_chars: int) -> str:
    # Drop scripts/styles quickly (best-effort).
    html = re.sub(r"(?is)<script[^>]*>.*?</script>", " ", html)
    html = re.sub(r"(?is)<style[^>]*>.*?</style>", " ", html)
    html = re.sub(r"(?is)<!--.*?-->", " ", html)
    # Reduce whitespace a bit.
    html = re.sub(r"\s+", " ", html).strip()
    if len(html) > max_chars:
        html = html[:max_chars]
    # Redact any emails so we don't leak them into the prompt.
    html = EMAIL_RE.sub("[REDACTED_EMAIL]", html)
    return html


def _fetch_html(url: str, *, timeout_s: int, user_agent: str) -> tuple[Optional[str], Optional[str]]:
    try:
        req = Request(url, headers={"User-Agent": user_agent})
        with urlopen(req, timeout=timeout_s) as resp:
            ctype = (resp.headers.get("Content-Type") or "").lower()
            if "text/html" not in ctype and "application/xhtml+xml" not in ctype and ctype:
                return None, f"non-html content-type: {ctype}"
            raw = resp.read()
            # naive decode; good enough for a best-effort pass
            html = raw.decode("utf-8", errors="ignore")
            return html, None
    except Exception as exc:
        return None, str(exc)


def _ollama_extract(ollama_mod: Any, *, model: str, website_url: str, html_stripped: str) -> dict[str, Any]:
    prompt = (
        "You are extracting contact information from a business website homepage HTML.\n"
        "The HTML has had any explicit emails redacted as [REDACTED_EMAIL].\n"
        "Your job is to infer or locate contact channels in the HTML, and suggest likely contact pages.\n\n"
        f"Website URL: {website_url}\n\n"
        "Return STRICT JSON with keys:\n"
        '{\n'
        '  "emails": string[],\n'
        '  "phones": string[],\n'
        '  "contact_pages": string[],\n'
        '  "socials": {"facebook": string|null, "instagram": string|null, "linkedin": string|null, "twitter": string|null},\n'
        '  "notes": string,\n'
        '  "confidence": "low"|"medium"|"high"\n'
        '}\n\n'
        "HTML:\n"
        f"{html_stripped}\n"
    )

    resp = ollama_mod.chat(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        format="json",
        options={"temperature": 0.2},
    )
    content = resp.get("message", {}).get("content", "") if isinstance(resp, dict) else ""
    try:
        data = json.loads(content)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    # Fallback: return raw text if model didn't respect JSON.
    return {"emails": [], "phones": [], "contact_pages": [], "socials": {}, "notes": str(content)[:5000], "confidence": "low"}


@dataclass
class EnrichedRow:
    maps_url: Optional[str]
    website: Optional[str]
    fetch_error: Optional[str]
    emails: str
    phones: str
    contact_pages: str
    facebook: Optional[str]
    instagram: Optional[str]
    linkedin: Optional[str]
    twitter: Optional[str]
    confidence: Optional[str]
    notes: Optional[str]


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Enrich leads by fetching websites and extracting contact info via local Ollama.")
    parser.add_argument("--in", dest="in_path", required=True, help="Input .xlsx path (e.g. spawn-radius-leads-v2.dedup.xlsx)")
    parser.add_argument("--sheet", default="leads_simple", help='Sheet to read from (default "leads_simple")')
    parser.add_argument("--out", default=None, help="Output .xlsx path (default: <input>.enriched.xlsx)")
    parser.add_argument("--ollama-model", default="qwen3.5:9b", help='Ollama model name (default "qwen3.5:9b")')
    parser.add_argument("--timeout-s", type=int, default=12, help="Website fetch timeout seconds (default 12)")
    parser.add_argument("--max-html-chars", type=int, default=140_000, help="Max HTML chars sent to Ollama (default 140k)")
    parser.add_argument("--sleep-ms", type=int, default=100, help="Small delay per lead to be polite (default 100ms)")
    parser.add_argument("--limit", type=int, default=0, help="Only process the first N website rows (default 0 = no limit)")
    args = parser.parse_args(argv)

    openpyxl = _require("openpyxl")
    ollama_mod = _require("ollama")
    from openpyxl import load_workbook  # type: ignore

    in_path = Path(args.in_path).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve() if args.out else in_path.with_name(in_path.stem + ".enriched" + in_path.suffix)

    wb = load_workbook(in_path)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f'Sheet "{args.sheet}" not found. Available: {wb.sheetnames}')
    ws = wb[args.sheet]

    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        raise SystemExit("No header row found.")
    headers = [str(h).strip() if h is not None else "" for h in headers]
    idx = {h: i for i, h in enumerate(headers) if h}

    if "website" not in idx and "maps_url" not in idx:
        raise SystemExit(f'Expected at least "website" or "maps_url" columns. Headers: {headers}')

    # Create/replace output sheet.
    out_sheet = "website_enriched"
    if out_sheet in wb.sheetnames:
        del wb[out_sheet]
    ws_out = wb.create_sheet(out_sheet)

    out_headers = [
        "maps_url",
        "website",
        "fetch_error",
        "emails",
        "phones",
        "contact_pages",
        "facebook",
        "instagram",
        "linkedin",
        "twitter",
        "confidence",
        "notes",
    ]
    ws_out.append(out_headers)

    ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"

    enriched_count = 0
    for r_i, row in enumerate(rows, start=2):
        if int(args.limit) > 0 and enriched_count >= int(args.limit):
            break
        website = _normalise_url(row[idx["website"]] if "website" in idx else None)
        maps_url = str(row[idx["maps_url"]]).strip() if "maps_url" in idx and row[idx["maps_url"]] else None
        if not website:
            continue

        html, fetch_err = _fetch_html(website, timeout_s=int(args.timeout_s), user_agent=ua)
        if not html:
            er = EnrichedRow(
                maps_url=maps_url,
                website=website,
                fetch_error=fetch_err or "fetch failed",
                emails="",
                phones="",
                contact_pages="",
                facebook=None,
                instagram=None,
                linkedin=None,
                twitter=None,
                confidence="low",
                notes=None,
            )
            ws_out.append([getattr(er, h) for h in out_headers])
            continue

        stripped = _strip_html_for_llm(html, max_chars=int(args.max_html_chars))
        data = _ollama_extract(ollama_mod, model=str(args.ollama_model), website_url=website, html_stripped=stripped)

        socials = data.get("socials") if isinstance(data, dict) else {}
        if not isinstance(socials, dict):
            socials = {}

        # Since emails were redacted, these will usually be empty unless the model inferred them.
        emails = data.get("emails", [])
        phones = data.get("phones", [])
        contact_pages = data.get("contact_pages", [])
        if not isinstance(emails, list):
            emails = []
        if not isinstance(phones, list):
            phones = []
        if not isinstance(contact_pages, list):
            contact_pages = []

        er = EnrichedRow(
            maps_url=maps_url,
            website=website,
            fetch_error=None,
            emails=", ".join([str(x) for x in emails if x]),
            phones=", ".join([str(x) for x in phones if x]),
            contact_pages=", ".join([str(x) for x in contact_pages if x]),
            facebook=socials.get("facebook"),
            instagram=socials.get("instagram"),
            linkedin=socials.get("linkedin"),
            twitter=socials.get("twitter"),
            confidence=data.get("confidence") if isinstance(data, dict) else None,
            notes=(data.get("notes") if isinstance(data, dict) else None),
        )
        ws_out.append([getattr(er, h) for h in out_headers])
        enriched_count += 1

        if int(args.sleep_ms) > 0:
            time.sleep(int(args.sleep_ms) / 1000.0)

        if enriched_count % 10 == 0:
            print(f"Enriched {enriched_count} website(s)...")

    wb.save(out_path)
    print(f"Wrote workbook with new sheet '{out_sheet}' to: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

