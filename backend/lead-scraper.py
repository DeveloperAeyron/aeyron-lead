#!/usr/bin/env python3
from __future__ import annotations

import argparse
import logging
import re
import sys
from urllib.parse import quote_plus
from dataclasses import dataclass, asdict, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

LOG = logging.getLogger("lead-scraper")


@dataclass(frozen=True)
class Lead:
    query: str
    location: str
    name: Optional[str] = None
    category: Optional[str] = None
    rating: Optional[str] = None
    reviews: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    plus_code: Optional[str] = None
    maps_url: Optional[str] = None
    scraped_at_iso: str = field(
        default_factory=lambda: datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    )


def _clean_whitespace(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    v = re.sub(r"\s+", " ", value).strip()
    # Google Maps sometimes prefixes fields with icon glyphs (private-use chars).
    v = re.sub(r"^[^A-Za-z0-9+(-]+", "", v).strip()
    return v or None


def _safe_inner_text(locator) -> Optional[str]:
    try:
        return _clean_whitespace(locator.inner_text(timeout=1500))
    except Exception:
        return None


def _extract_from_aria_label(page, label_contains: str) -> Optional[str]:
    """
    Google Maps place details expose fields as buttons with aria-labels like:
    - "Address: 123 Example St"
    - "Phone: +1 302-555-0100"
    - "Website: example.com"
    - "Plus code: ABCD+12 City, State"
    """
    candidates = page.locator(f'[role="button"][aria-label*="{label_contains}"], [aria-label*="{label_contains}"][role="link"]')
    if candidates.count() < 1:
        return None
    aria = candidates.first.get_attribute("aria-label") or ""
    aria = _clean_whitespace(aria)
    if not aria:
        return None
    parts = aria.split(":", 1)
    if len(parts) == 2:
        return _clean_whitespace(parts[1])
    return aria


def _extract_from_data_item_id(page, key: str) -> Optional[str]:
    """
    Alternate extraction path used by Google Maps place panels.
    Common keys include: address, phone, authority (website), oloc (plus code).
    """
    loc = page.locator(f'button[data-item-id*="{key}"], a[data-item-id*="{key}"]').first
    text = _safe_inner_text(loc)
    return _clean_whitespace(text)


def _first_attr(page, selector: str, attr: str) -> Optional[str]:
    loc = page.locator(selector).first
    try:
        val = loc.get_attribute(attr, timeout=1500)
        return _clean_whitespace(val)
    except Exception:
        return None


def _first_attr_with_digits(page, selector: str, attr: str) -> Optional[str]:
    loc = page.locator(selector)
    try:
        n = min(loc.count(), 8)
    except Exception:
        return None
    for i in range(n):
        try:
            val = _clean_whitespace(loc.nth(i).get_attribute(attr, timeout=1500))
        except Exception:
            continue
        if val and re.search(r"\d", val):
            return val
    return None


_EMAIL_RE = re.compile(r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,}\b")


def _extract_email_from_maps_page(page) -> Optional[str]:
    """
    Emails are not consistently exposed in the Maps panel, but occasionally appear
    in visible text (e.g. business description / review snippets) or in mailto: links.
    This is a best-effort extraction from the current place page only.
    """
    mailto = page.locator('a[href^="mailto:"]').first
    try:
        href = mailto.get_attribute("href", timeout=800)
        if href:
            m = _EMAIL_RE.search(href)
            if m:
                return _clean_whitespace(m.group(0))
    except Exception:
        pass

    try:
        body_text = page.locator("body").inner_text(timeout=1500)
        m = _EMAIL_RE.search(body_text or "")
        if m:
            return _clean_whitespace(m.group(0))
    except Exception:
        return None

    return None

def _wait_for_results(page, timeout_ms: int) -> None:
    # Either a list of results appears, or Google takes you straight to a place page.
    page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    page.wait_for_timeout(250)
    try:
        page.wait_for_selector(
            'a[href*="/maps/place/"], div[role="feed"], div[role="main"] h1',
            timeout=timeout_ms,
        )
    except PlaywrightTimeoutError:
        # If we are already on a place page, this is fine.
        pass


def _click_first_result_if_present(page) -> None:
    """
    Prefer opening a specific place page from a search results list.
    Google Maps DOM shifts frequently, so we try a small set of selectors.
    """
    candidate_selectors = [
        # Common card link class in results list.
        'a.hfpxzc[href*="/maps/place/"]',
        # Generic place links (sometimes only these exist).
        'a[href*="/maps/place/"][aria-label]',
        'a[href*="/maps/place/"]',
    ]
    for sel in candidate_selectors:
        loc = page.locator(sel).first
        try:
            if loc.is_visible(timeout=1500):
                loc.click()
                return
        except Exception:
            continue


def _normalise_maps_url(url: str) -> str:
    # Keep it simple for dedupe: strip URL fragments.
    return url.split("#", 1)[0]


def _collect_place_urls(page, limit: int, timeout_ms: int) -> list[str]:
    """
    Collect up to `limit` place URLs from the search results list by scrolling.
    """
    selectors = [
        'a.hfpxzc[href*="/maps/place/"]',
        'a[href*="/maps/place/"][aria-label]',
        'a[href*="/maps/place/"]',
    ]

    seen: set[str] = set()
    urls: list[str] = []

    def grab_once() -> None:
        nonlocal urls
        for sel in selectors:
            loc = page.locator(sel)
            try:
                n = min(loc.count(), 50)
            except Exception:
                continue
            for i in range(n):
                href = loc.nth(i).get_attribute("href")
                if not href:
                    continue
                if href.startswith("/"):
                    href = "https://www.google.com" + href
                href = _normalise_maps_url(href)
                if "/maps/place/" not in href:
                    continue
                if href in seen:
                    continue
                seen.add(href)
                urls.append(href)
                if len(urls) >= limit:
                    return

    grab_once()
    if len(urls) >= limit:
        return urls[:limit]

    # Scroll the results panel (when present) to load more.
    results_feed = page.locator('div[role="feed"]').first
    for _ in range(25):
        if len(urls) >= limit:
            break
        try:
            if results_feed.is_visible(timeout=500):
                results_feed.evaluate("(el) => { el.scrollBy(0, el.scrollHeight); }")
            else:
                page.mouse.wheel(0, 1400)
        except Exception:
            page.mouse.wheel(0, 1400)
        page.wait_for_timeout(500)
        grab_once()

    return urls[:limit]


def _scrape_place_page(page, *, query: str, location: str, timeout_ms: int) -> Lead:
    page.set_default_timeout(timeout_ms)
    page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    page.wait_for_timeout(750)

    name = None
    for loc in [
        page.locator("h1.DUwDvf").first,
        page.locator('div[role="main"] h1').first,
        page.locator("h1").first,
    ]:
        candidate = _safe_inner_text(loc)
        if candidate and candidate.lower() != "results":
            name = candidate
            break

    if not name:
        try:
            title = _clean_whitespace(page.title())
            if title and " - Google Maps" in title:
                name = _clean_whitespace(title.replace(" - Google Maps", ""))
        except Exception:
            pass

    category = _safe_inner_text(
        page.locator('button[jsaction*="pane.rating.category"], button[aria-label*="Category"]').first
    ) or _safe_inner_text(page.locator('button[jsaction*="pane.rating.category"]').first)

    rating = (
        _safe_inner_text(page.locator('div.F7nice span[aria-hidden="true"]').first)
        or _first_attr(page, 'span[aria-label*="stars"]', "aria-label")
        or _safe_inner_text(page.locator('span[aria-label*="stars"]').first)
    )
    if rating:
        m = re.search(r"([0-9.]+)\s*stars", rating, re.I)
        rating = m.group(1) if m else _clean_whitespace(rating)

    reviews = (
        _first_attr_with_digits(page, 'button[aria-label*="reviews"]', "aria-label")
        or _extract_from_aria_label(page, "Reviews")
        or _extract_from_aria_label(page, "reviews")
    )
    if reviews:
        m = re.search(r"(\d[\d,]*)", reviews)
        reviews = m.group(1) if m else reviews

    address = _extract_from_aria_label(page, "Address") or _extract_from_data_item_id(page, "address")
    phone = _extract_from_aria_label(page, "Phone") or _extract_from_data_item_id(page, "phone")
    website = _extract_from_aria_label(page, "Website") or _extract_from_data_item_id(page, "authority")
    email = _extract_email_from_maps_page(page)
    plus_code = _extract_from_aria_label(page, "Plus code") or _extract_from_data_item_id(page, "oloc")

    return Lead(
        query=query,
        location=location,
        name=name,
        category=category,
        rating=rating,
        reviews=reviews,
        address=address,
        phone=phone,
        website=website,
        email=email,
        plus_code=plus_code,
        maps_url=page.url,
    )


def scrape_one(
    *,
    query: str,
    location: str,
    headless: bool,
    slow_mo_ms: int,
    timeout_ms: int,
) -> Lead:
    search_text = f"{query} in {location}".strip()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms)
        context = browser.new_context(
            locale="en-GB",
            viewport={"width": 1400, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()
        page.set_default_timeout(timeout_ms)

        search_url = f"https://www.google.com/maps/search/{quote_plus(search_text)}"
        page.goto(search_url, wait_until="domcontentloaded", timeout=timeout_ms)

        _wait_for_results(page, timeout_ms)

        # Prefer opening the first result, if a list is shown.
        _click_first_result_if_present(page)

        # Wait for place details panel content.
        lead = _scrape_place_page(page, query=query, location=location, timeout_ms=timeout_ms)

        context.close()
        browser.close()

        return lead


def write_xlsx(leads: Iterable[Lead], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "leads"

    leads_list = list(leads)
    if not leads_list:
        leads_list = [Lead(query="", location="")]

    headers = list(asdict(leads_list[0]).keys())
    ws.append(headers)
    for lead in leads_list:
        row = asdict(lead)
        ws.append([row.get(h) for h in headers])

    # Basic sizing.
    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx).value
            max_len = max(max_len, len(str(v or "")))
        ws.column_dimensions[col].width = min(max(12, max_len + 2), 60)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def scrape_many(
    *,
    query: str,
    location: str,
    limit: int,
    headless: bool,
    slow_mo_ms: int,
    timeout_ms: int,
) -> list[Lead]:
    search_text = f"{query} in {location}".strip()
    if limit < 1:
        return []

    started = datetime.now(timezone.utc)
    LOG.info('Starting scrape: query="%s" location="%s" limit=%s headless=%s', query, location, limit, headless)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms)
        context = browser.new_context(
            locale="en-GB",
            viewport={"width": 1400, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()
        page.set_default_timeout(timeout_ms)

        search_url = f"https://www.google.com/maps/search/{quote_plus(search_text)}"
        page.goto(search_url, wait_until="domcontentloaded", timeout=timeout_ms)
        _wait_for_results(page, timeout_ms)

        place_urls = _collect_place_urls(page, limit=limit, timeout_ms=timeout_ms)
        LOG.info("Collected %s place URL(s) from results", len(place_urls))
        leads: list[Lead] = []
        seen: set[str] = set()

        for idx, url in enumerate(place_urls, start=1):
            if url in seen:
                continue
            seen.add(url)
            try:
                elapsed_s = (datetime.now(timezone.utc) - started).total_seconds()
                LOG.info("Scraping %s/%s (elapsed %.1fs)", idx, len(place_urls), elapsed_s)
                page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
                leads.append(_scrape_place_page(page, query=query, location=location, timeout_ms=timeout_ms))
            except Exception:
                LOG.exception("Failed to scrape place page")
                leads.append(Lead(query=query, location=location, maps_url=url))

        context.close()
        browser.close()

        total_s = (datetime.now(timezone.utc) - started).total_seconds()
        LOG.info("Finished scrape: %s lead(s) in %.1fs", len(leads), total_s)
        return leads


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Scrape Google Maps leads into an .xlsx file.")
    parser.add_argument("--query", default="car wash", help='Search term, e.g. "car wash"')
    parser.add_argument("--location", default="Delaware", help='Location, e.g. "Delaware"')
    parser.add_argument("--limit", type=int, default=10, help="Number of shops to collect (default 10)")
    parser.add_argument("--headless", action="store_true", help="Run headless (default is headed)")
    parser.add_argument("--slowmo-ms", type=int, default=50, help="Slow motion delay per action (ms)")
    parser.add_argument("--timeout-ms", type=int, default=60_000, help="Playwright timeout (ms)")
    parser.add_argument("--out", default="lead.xlsx", help="Output .xlsx path")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    leads = scrape_many(
        query=args.query,
        location=args.location,
        limit=args.limit,
        headless=args.headless,
        slow_mo_ms=args.slowmo_ms,
        timeout_ms=args.timeout_ms,
    )

    out_path = Path(args.out).expanduser().resolve()
    write_xlsx(leads, out_path)

    print(f"Wrote {len(leads)} lead(s) to: {out_path}")
    if leads:
        print(f"First: {leads[0].name or '(missing)'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

