#!/usr/bin/env python3
"""High-performance spawn-radius scraper v2: async workers, resource blocking, two-phase pipeline."""
from __future__ import annotations

import argparse, asyncio, json, logging, math, random, re, sys
from collections import deque
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Deque, Iterable, Optional
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright

LOG = logging.getLogger("spawn-radius-v2")

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
]

# ── Data Model (same contract as v1 for SSE compatibility) ─────────────
@dataclass(frozen=True)
class Lead:
    root_query: str
    root_location: str
    root_seed_name: Optional[str] = None
    root_seed_lat: Optional[float] = None
    root_seed_lng: Optional[float] = None
    spawn_depth: int = 0
    parent_seed_name: Optional[str] = None
    parent_seed_lat: Optional[float] = None
    parent_seed_lng: Optional[float] = None
    radius_km: float = 1.0
    name: Optional[str] = None
    rating: Optional[str] = None
    reviews: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    plus_code: Optional[str] = None
    maps_url: Optional[str] = None
    place_lat: Optional[float] = None
    place_lng: Optional[float] = None
    distance_km_from_parent: Optional[float] = None
    scraped_at_iso: str = field(
        default_factory=lambda: datetime.now(timezone.utc)
        .replace(microsecond=0).isoformat().replace("+00:00", "Z")
    )

_EMAIL_RE = re.compile(r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,}\b")

# ── Helpers ────────────────────────────────────────────────────────────
def _clean(value: Optional[str]) -> Optional[str]:
    if value is None: return None
    v = re.sub(r"\s+", " ", value).strip()
    v = re.sub(r"^[^A-Za-z0-9+(-]+", "", v).strip()
    return v or None

def _normalise_url(url: str) -> str:
    return url.split("#", 1)[0]

def _parse_coords(url: str) -> tuple[Optional[float], Optional[float]]:
    m = re.search(r"/@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?),", url)
    if m: return float(m.group(1)), float(m.group(2))
    m = re.search(r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)", url)
    if m: return float(m.group(1)), float(m.group(2))
    return None, None

def _haversine_km(lat1, lng1, lat2, lng2) -> float:
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lng2 - lng1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * r * math.atan2(math.sqrt(a), math.sqrt(1 - a))

# ── Resource Blocker (40-60% speedup) ──────────────────────────────────
BLOCKED_TYPES = {"image", "font", "media", "stylesheet"}

async def _block_resources(route):
    if route.request.resource_type in BLOCKED_TYPES:
        await route.abort()
    else:
        await route.continue_()

# ── Async Playwright Helpers ───────────────────────────────────────────
async def _txt(locator) -> Optional[str]:
    try: return _clean(await locator.inner_text(timeout=1500))
    except Exception: return None

async def _attr(page, sel, attr) -> Optional[str]:
    try: return _clean(await page.locator(sel).first.get_attribute(attr, timeout=1500))
    except Exception: return None

async def _aria(page, label) -> Optional[str]:
    loc = page.locator(f'[role="button"][aria-label*="{label}"], [aria-label*="{label}"][role="link"]')
    try:
        if await loc.count() < 1: return None
        a = _clean(await loc.first.get_attribute("aria-label") or "")
        if not a: return None
        parts = a.split(":", 1)
        return _clean(parts[1]) if len(parts) == 2 else a
    except Exception: return None

async def _data_item(page, key) -> Optional[str]:
    return _clean(await _txt(page.locator(f'button[data-item-id*="{key}"], a[data-item-id*="{key}"]').first))

async def _attr_digits(page, sel, attr) -> Optional[str]:
    loc = page.locator(sel)
    try: n = min(await loc.count(), 10)
    except Exception: return None
    for i in range(n):
        try: val = _clean(await loc.nth(i).get_attribute(attr, timeout=1500))
        except Exception: continue
        if val and re.search(r"\d", val): return val
    return None

async def _email(page) -> Optional[str]:
    try:
        href = await page.locator('a[href^="mailto:"]').first.get_attribute("href", timeout=800)
        if href:
            m = _EMAIL_RE.search(href)
            if m: return _clean(m.group(0))
    except Exception: pass
    try:
        body = await page.locator("body").inner_text(timeout=1500)
        m = _EMAIL_RE.search(body or "")
        if m: return _clean(m.group(0))
    except Exception: pass
    return None

# ── Smart Wait (replaces blind sleeps) ─────────────────────────────────
async def _smart_wait_results(page, timeout_ms: int = 12000) -> None:
    await page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    try:
        await page.wait_for_selector(
            'a[href*="/maps/place/"], div[role="feed"], div[role="main"] h1',
            timeout=timeout_ms,
        )
    except PlaywrightTimeoutError:
        pass

async def _smart_wait_place(page, timeout_ms: int = 8000) -> None:
    await page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    try:
        await page.wait_for_selector('h1.DUwDvf, div[role="main"] h1', state="visible", timeout=timeout_ms)
    except PlaywrightTimeoutError:
        pass

# ── URL Harvesting ─────────────────────────────────────────────────────
async def _harvest_urls(page, limit: int) -> list[str]:
    seen, urls = set(), []
    async def grab():
        for sel in ['a.hfpxzc[href*="/maps/place/"]', 'a[href*="/maps/place/"]']:
            loc = page.locator(sel)
            try: n = min(await loc.count(), 200)
            except Exception: continue
            for i in range(n):
                try: href = await loc.nth(i).get_attribute("href")
                except Exception: continue
                if not href: continue
                if href.startswith("/"): href = "https://www.google.com" + href
                href = _normalise_url(href)
                if "/maps/place/" not in href or href in seen: continue
                seen.add(href); urls.append(href)
                if len(urls) >= limit: return
    await grab()
    if len(urls) >= limit: return urls[:limit]
    feed = page.locator('div[role="feed"]').first
    stale = 0
    for _ in range(300):
        if len(urls) >= limit or stale >= 8: break
        before = len(urls)
        try:
            if await feed.is_visible(timeout=500):
                await feed.evaluate("el => el.scrollBy(0, el.scrollHeight)")
            else:
                await page.mouse.wheel(0, 1400)
        except Exception:
            try: await page.mouse.wheel(0, 1400)
            except Exception: pass
        await page.wait_for_timeout(400)
        await grab()
        stale = stale + 1 if len(urls) == before else 0
    return urls[:limit]

# ── Place Detail Scraper ───────────────────────────────────────────────
async def _scrape_place(page, *, root_query, root_location, root_seed_name,
                        root_seed_lat, root_seed_lng, depth, parent_seed_name,
                        parent_seed_lat, parent_seed_lng, radius_km) -> Lead:
    await _smart_wait_place(page)
    name = None
    for loc in [page.locator("h1.DUwDvf").first, page.locator('div[role="main"] h1').first, page.locator("h1").first]:
        c = await _txt(loc)
        if c and c.lower() != "results": name = c; break
    if not name:
        try:
            t = _clean(await page.title())
            if t and " - Google Maps" in t: name = _clean(t.replace(" - Google Maps", ""))
        except Exception: pass
    rating = (await _txt(page.locator('div.F7nice span[aria-hidden="true"]').first)
              or await _attr(page, 'span[aria-label*="stars"]', "aria-label")
              or await _txt(page.locator('span[aria-label*="stars"]').first))
    if rating:
        m = re.search(r"([0-9.]+)\s*stars", rating, re.I)
        rating = m.group(1) if m else _clean(rating)
    reviews = (await _attr_digits(page, 'button[aria-label*="reviews"]', "aria-label")
               or await _aria(page, "Reviews") or await _aria(page, "reviews"))
    if reviews:
        m = re.search(r"(\d[\d,]*)", reviews)
        reviews = m.group(1) if m else reviews
    address = await _aria(page, "Address") or await _data_item(page, "address")
    phone = await _aria(page, "Phone") or await _data_item(page, "phone")
    website = await _aria(page, "Website") or await _data_item(page, "authority")
    plus_code = await _aria(page, "Plus code") or await _data_item(page, "oloc")
    em = await _email(page)
    lat, lng = _parse_coords(page.url)
    dist = _haversine_km(parent_seed_lat, parent_seed_lng, lat, lng) if lat and lng else None
    return Lead(root_query=root_query, root_location=root_location,
                root_seed_name=root_seed_name, root_seed_lat=root_seed_lat,
                root_seed_lng=root_seed_lng, spawn_depth=depth,
                parent_seed_name=parent_seed_name, parent_seed_lat=parent_seed_lat,
                parent_seed_lng=parent_seed_lng, radius_km=radius_km,
                name=name, rating=rating, reviews=reviews, address=address,
                phone=phone, website=website, email=em, plus_code=plus_code,
                maps_url=page.url, place_lat=lat, place_lng=lng,
                distance_km_from_parent=dist)

# ── Page Pool ──────────────────────────────────────────────────────────
class PagePool:
    def __init__(self, context, size):
        self._sem = asyncio.Semaphore(size)
        self._q: asyncio.Queue = asyncio.Queue()
        self._ctx = context
    async def init(self):
        for _ in range(self._sem._value):
            await self._q.put(await self._ctx.new_page())
    async def acquire(self):
        await self._sem.acquire()
        return await self._q.get()
    async def release(self, page):
        try: await page.goto("about:blank", wait_until="commit", timeout=3000)
        except Exception:
            try: page = await self._ctx.new_page()
            except Exception: pass
        await self._q.put(page); self._sem.release()
    async def close_all(self):
        while not self._q.empty():
            try: await (await self._q.get()).close()
            except Exception: pass

# ── XLSX Writer ────────────────────────────────────────────────────────
def write_xlsx(leads: Iterable[Lead], out_path: Path) -> None:
    wb = Workbook(); ws = wb.active; ws.title = "leads"
    ll = list(leads)
    if not ll: ll = [Lead(root_query="", root_location="")]
    headers = list(asdict(ll[0]).keys()); ws.append(headers)
    for lead in ll:
        row = asdict(lead); ws.append([row.get(h) for h in headers])
    for i, h in enumerate(headers, 1):
        col = get_column_letter(i)
        mx = max(len(h), *(len(str(ws.cell(r, i).value or "")) for r in range(2, ws.max_row+1)))
        ws.column_dimensions[col].width = min(max(12, mx+2), 70)
    out_path.parent.mkdir(parents=True, exist_ok=True); wb.save(out_path)

# ══════════════════════════════════════════════════════════════════════
#  MAIN ENGINE — Two-Phase Pipeline with Async Worker Pool
# ══════════════════════════════════════════════════════════════════════
async def spawn_scrape_v2(
    *, root_query, root_location, nearby_query, radius_km, root_count,
    root_skip, per_seed_candidates, per_seed_keep_cap, max_depth, max_total,
    checkpoint_path, checkpoint_every, headless, slow_mo_ms, timeout_ms,
    zoom, workers, json_stream=False,
) -> list[Lead]:
    started = datetime.now(timezone.utc)
    LOG.info("v2 scrape: query=%r loc=%r radius=%.1fkm workers=%s max=%s depth=%s",
             root_query, root_location, radius_km, workers, max_total, max_depth)

    search_text = f"{root_query} in {root_location}".strip()
    kept, visited = [], set()
    seen_leads: set[str] = set()  # dedup keys for leads
    lock = asyncio.Lock()

    def _lead_key(lead: Lead) -> str:
        """Generate a dedup key from maps_url or name+address+phone fallback."""
        url = _normalise_url(lead.maps_url or "").strip()
        if url:
            return f"url:{url}"
        return f"nap:{(lead.name or '').strip().lower()}|{(lead.address or '').strip().lower()}|{(lead.phone or '').strip()}"

    def _checkpoint():
        if not checkpoint_path or checkpoint_every <= 0: return
        if len(kept) % checkpoint_every != 0: return
        try: write_xlsx(kept, checkpoint_path); LOG.info("Checkpoint: %s leads", len(kept))
        except Exception: LOG.exception("Checkpoint failed")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=headless, slow_mo=slow_mo_ms,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        ctx = await browser.new_context(
            locale="en-GB", viewport={"width": 1400, "height": 900},
            user_agent=random.choice(USER_AGENTS),
        )
        pool = PagePool(ctx, size=max(2, workers))
        await pool.init()

        # ── Phase 1: Collect root URLs ─────────────────────────────────
        pg = await pool.acquire()
        try:
            url = f"https://www.google.com/maps/search/{quote_plus(search_text)}"
            await pg.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            await _smart_wait_results(pg, timeout_ms)
            root_urls = await _harvest_urls(pg, limit=root_skip + root_count)
        except Exception:
            LOG.exception("Failed to collect root seeds"); root_urls = []
        finally:
            await pool.release(pg)

        if not root_urls:
            await pool.close_all(); await ctx.close(); await browser.close()
            raise RuntimeError("No place URLs found.")
        root_urls = root_urls[root_skip:root_skip + root_count]
        LOG.info("Collected %s root URLs", len(root_urls))

        # ── Phase 2: Resolve roots concurrently ────────────────────────
        resolved = []
        async def resolve_root(idx, rurl):
            rurl = _normalise_url(rurl)
            async with lock:
                if rurl in visited: return
            pg = await pool.acquire()
            try:
                await pg.goto(rurl, wait_until="domcontentloaded", timeout=timeout_ms)
                await _smart_wait_place(pg)
                name = (await _txt(pg.locator("h1.DUwDvf").first)
                        or await _txt(pg.locator("h1").first) or "(unknown)")
                lat, lng = _parse_coords(pg.url)
                async with lock:
                    visited.add(_normalise_url(pg.url))
                if lat is None or lng is None:
                    LOG.info("Skipping root %r (no coords)", name); return
                async with lock:
                    resolved.append((name, rurl, lat, lng))
                LOG.info("Root %s/%s: %r (%.4f, %.4f)", idx, len(root_urls), name, lat, lng)
            except Exception:
                LOG.exception("Failed root %s", rurl)
            finally:
                await pool.release(pg)

        await asyncio.gather(*[resolve_root(i, u) for i, u in enumerate(root_urls, 1)])
        LOG.info("Resolved %s roots concurrently", len(resolved))

        # ── Phase 3: BFS expansion with worker pool ────────────────────
        seed_q: asyncio.Queue = asyncio.Queue()
        stop = asyncio.Event()

        for name, url, lat, lng in resolved:
            await seed_q.put((name, lat, lng, 0, name, lat, lng))

        async def seed_worker(wid):
            while not stop.is_set():
                try: seed = await asyncio.wait_for(seed_q.get(), timeout=5.0)
                except asyncio.TimeoutError:
                    if seed_q.empty(): break
                    continue
                pname, plat, plng, depth, rsname, rslat, rslng = seed
                if depth >= max_depth:
                    seed_q.task_done(); continue
                async with lock:
                    if len(kept) >= max_total:
                        stop.set(); seed_q.task_done(); break

                pg = await pool.acquire()
                try:
                    nurl = f"https://www.google.com/maps/search/{quote_plus(nearby_query)}/@{plat},{plng},{zoom}z"
                    await pg.goto(nurl, wait_until="domcontentloaded", timeout=timeout_ms)
                    await _smart_wait_results(pg, timeout_ms)
                    cands = await _harvest_urls(pg, limit=per_seed_candidates)

                    within = []
                    for cu in cands:
                        cu = _normalise_url(cu)
                        async with lock:
                            if cu in visited: continue
                        lt, ln = _parse_coords(cu)
                        if lt is None or ln is None: continue
                        d = _haversine_km(plat, plng, lt, ln)
                        if 0 < d <= radius_km: within.append((cu, d, lt, ln))
                    within.sort(key=lambda t: t[1])
                    within = within[:per_seed_keep_cap]

                    scraped = 0
                    for cu, dist, lt, ln in within:
                        if stop.is_set(): break
                        async with lock:
                            if len(kept) >= max_total: stop.set(); break
                            if cu in visited: continue
                        try:
                            await pg.goto(cu, wait_until="domcontentloaded", timeout=timeout_ms)
                            async with lock: visited.add(_normalise_url(pg.url))
                            lead = await _scrape_place(
                                pg, root_query=root_query, root_location=root_location,
                                root_seed_name=rsname, root_seed_lat=rslat, root_seed_lng=rslng,
                                depth=depth+1, parent_seed_name=pname,
                                parent_seed_lat=plat, parent_seed_lng=plng, radius_km=radius_km,
                            )
                        except Exception:
                            LOG.exception("Worker %s: scrape failed", wid); continue
                        if lead.distance_km_from_parent is None or lead.distance_km_from_parent > radius_km:
                            continue
                        async with lock:
                            key = _lead_key(lead)
                            if key in seen_leads: continue
                            seen_leads.add(key)
                            kept.append(lead); scraped += 1
                        if json_stream:
                            print(json.dumps(asdict(lead)), flush=True)
                        _checkpoint()
                        if lead.place_lat and lead.place_lng and lead.name:
                            await seed_q.put((lead.name, lead.place_lat, lead.place_lng,
                                              depth+1, rsname, rslat, rslng))
                        if scraped >= per_seed_keep_cap: break
                except Exception:
                    LOG.exception("Worker %s: seed expansion failed", wid)
                finally:
                    await pool.release(pg)
                    seed_q.task_done()

        tasks = [asyncio.create_task(seed_worker(i)) for i in range(max(1, workers))]
        await asyncio.gather(*tasks)
        await pool.close_all(); await ctx.close(); await browser.close()

    elapsed = (datetime.now(timezone.utc) - started).total_seconds()
    LOG.info("Done. %s leads in %.1fs (%.1f leads/min)", len(kept), elapsed,
             len(kept) / max(elapsed/60, 0.01))
    return kept[:max_total]


def main(argv):
    ap = argparse.ArgumentParser(description="Spawn-radius scraper v2: high-performance async pipeline")
    ap.add_argument("--query", default="car wash")
    ap.add_argument("--location", default="Delaware")
    ap.add_argument("--nearby-query", default=None)
    ap.add_argument("--radius-km", type=float, default=1.0)
    ap.add_argument("--max-total", type=int, default=50)
    ap.add_argument("--root-count", type=int, default=50)
    ap.add_argument("--root-skip", type=int, default=0)
    ap.add_argument("--max-depth", type=int, default=3)
    ap.add_argument("--per-seed-candidates", type=int, default=80)
    ap.add_argument("--per-seed-keep-cap", type=int, default=10)
    ap.add_argument("--zoom", type=int, default=15)
    ap.add_argument("--checkpoint-every", type=int, default=5)
    ap.add_argument("--headless", action="store_true", help="Run headless (default headed)")
    ap.add_argument("--slowmo-ms", type=int, default=0)
    ap.add_argument("--timeout-ms", type=int, default=15000, help="Tight timeout (default 15s)")
    ap.add_argument("--json-stream", action="store_true")
    ap.add_argument("--workers", type=int, default=10, help="Concurrent scraping workers (default 10)")
    ap.add_argument("--out", default="spawn-radius-leads-v2.xlsx")
    args = ap.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
    nq = args.nearby_query or args.query
    out = Path(args.out).expanduser().resolve()
    leads = asyncio.run(spawn_scrape_v2(
        root_query=args.query, root_location=args.location, nearby_query=nq,
        radius_km=args.radius_km, root_count=args.root_count, root_skip=args.root_skip,
        per_seed_candidates=args.per_seed_candidates, per_seed_keep_cap=args.per_seed_keep_cap,
        max_depth=args.max_depth, max_total=args.max_total, checkpoint_path=out,
        checkpoint_every=args.checkpoint_every, headless=args.headless,
        slow_mo_ms=args.slowmo_ms, timeout_ms=args.timeout_ms, zoom=args.zoom,
        workers=args.workers, json_stream=args.json_stream,
    ))
    write_xlsx(leads, out)
    print(f"Wrote {len(leads)} lead(s) to: {out}")
    return 0

if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
