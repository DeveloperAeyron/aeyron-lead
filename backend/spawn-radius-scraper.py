#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import logging
import math
import re
import sys
from collections import deque
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Deque, Iterable, Optional
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

LOG = logging.getLogger("spawn-radius-scraper")


@dataclass(frozen=True)
class Lead:
    root_query: str
    root_location: str
    root_seed_name: Optional[str] = None
    root_seed_lat: Optional[float] = None
    root_seed_lng: Optional[float] = None
    spawn_depth: int = 0
    parent_seed_name: Optional[str] = None
    parent_seed_lat: Optional[float] = None
    parent_seed_lng: Optional[float] = None
    radius_km: float = 1.0
    name: Optional[str] = None
    rating: Optional[str] = None
    reviews: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None
    plus_code: Optional[str] = None
    maps_url: Optional[str] = None
    place_lat: Optional[float] = None
    place_lng: Optional[float] = None
    distance_km_from_parent: Optional[float] = None
    scraped_at_iso: str = field(
        default_factory=lambda: datetime.now(timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


_EMAIL_RE = re.compile(r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,}\b")


def _clean_whitespace(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    v = re.sub(r"\s+", " ", value).strip()
    v = re.sub(r"^[^A-Za-z0-9+(-]+", "", v).strip()
    return v or None


def _safe_inner_text(locator) -> Optional[str]:
    try:
        return _clean_whitespace(locator.inner_text(timeout=1500))
    except Exception:
        return None


def _wait_for_results(page, timeout_ms: int) -> None:
    page.wait_for_load_state("domcontentloaded", timeout=timeout_ms)
    page.wait_for_timeout(250)
    try:
        page.wait_for_selector('a[href*="/maps/place/"], div[role="feed"], div[role="main"] h1', timeout=timeout_ms)
    except PlaywrightTimeoutError:
        pass


def _normalise_maps_url(url: str) -> str:
    return url.split("#", 1)[0]


def _parse_coords_from_url(url: str) -> tuple[Optional[float], Optional[float]]:
    m = re.search(r"/@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?),", url)
    if m:
        return float(m.group(1)), float(m.group(2))
    m = re.search(r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)", url)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def _haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    r = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _extract_from_aria_label(page, label_contains: str) -> Optional[str]:
    candidates = page.locator(
        f'[role="button"][aria-label*="{label_contains}"], [aria-label*="{label_contains}"][role="link"]'
    )
    try:
        if candidates.count() < 1:
            return None
    except Exception:
        return None
    aria = candidates.first.get_attribute("aria-label") or ""
    aria = _clean_whitespace(aria)
    if not aria:
        return None
    parts = aria.split(":", 1)
    if len(parts) == 2:
        return _clean_whitespace(parts[1])
    return aria


def _extract_from_data_item_id(page, key: str) -> Optional[str]:
    loc = page.locator(f'button[data-item-id*="{key}"], a[data-item-id*="{key}"]').first
    return _clean_whitespace(_safe_inner_text(loc))


def _first_attr(page, selector: str, attr: str) -> Optional[str]:
    loc = page.locator(selector).first
    try:
        val = loc.get_attribute(attr, timeout=1500)
        return _clean_whitespace(val)
    except Exception:
        return None


def _first_attr_with_digits(page, selector: str, attr: str) -> Optional[str]:
    loc = page.locator(selector)
    try:
        n = min(loc.count(), 10)
    except Exception:
        return None
    for i in range(n):
        try:
            val = _clean_whitespace(loc.nth(i).get_attribute(attr, timeout=1500))
        except Exception:
            continue
        if val and re.search(r"\d", val):
            return val
    return None


def _extract_email_from_maps_page(page) -> Optional[str]:
    mailto = page.locator('a[href^="mailto:"]').first
    try:
        href = mailto.get_attribute("href", timeout=800)
        if href:
            m = _EMAIL_RE.search(href)
            if m:
                return _clean_whitespace(m.group(0))
    except Exception:
        pass

    try:
        body_text = page.locator("body").inner_text(timeout=1500)
        m = _EMAIL_RE.search(body_text or "")
        if m:
            return _clean_whitespace(m.group(0))
    except Exception:
        return None

    return None


def _collect_place_urls(page, limit: int) -> list[str]:
    selectors = [
        'a.hfpxzc[href*="/maps/place/"]',
        'a[href*="/maps/place/"][aria-label]',
        'a[href*="/maps/place/"]',
    ]

    seen: set[str] = set()
    urls: list[str] = []

    def grab_once() -> None:
        for sel in selectors:
            loc = page.locator(sel)
            try:
                n = min(loc.count(), 120)
            except Exception:
                continue
            for i in range(n):
                href = loc.nth(i).get_attribute("href")
                if not href:
                    continue
                if href.startswith("/"):
                    href = "https://www.google.com" + href
                href = _normalise_maps_url(href)
                if "/maps/place/" not in href:
                    continue
                if href in seen:
                    continue
                seen.add(href)
                urls.append(href)
                if len(urls) >= limit:
                    return

    grab_once()
    if len(urls) >= limit:
        return urls[:limit]

    results_feed = page.locator('div[role="feed"]').first
    no_new = 0
    max_no_new = 8

    for _ in range(300):  # safety cap
        if len(urls) >= limit or no_new >= max_no_new:
            break
        before = len(urls)
        try:
            if results_feed.is_visible(timeout=500):
                results_feed.evaluate("(el) => { el.scrollBy(0, el.scrollHeight); }")
            else:
                page.mouse.wheel(0, 1400)
        except Exception:
            page.mouse.wheel(0, 1400)
        page.wait_for_timeout(650)
        grab_once()
        no_new = no_new + 1 if len(urls) == before else 0

    return urls[:limit]


def _scrape_place_page(
    page,
    *,
    root_query: str,
    root_location: str,
    root_seed_name: str,
    root_seed_lat: float,
    root_seed_lng: float,
    depth: int,
    parent_seed_name: str,
    parent_seed_lat: float,
    parent_seed_lng: float,
    radius_km: float,
) -> Lead:
    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(550)

    name = None
    for loc in [page.locator("h1.DUwDvf").first, page.locator('div[role="main"] h1').first, page.locator("h1").first]:
        candidate = _safe_inner_text(loc)
        if candidate and candidate.lower() != "results":
            name = candidate
            break
    if not name:
        try:
            title = _clean_whitespace(page.title())
            if title and " - Google Maps" in title:
                name = _clean_whitespace(title.replace(" - Google Maps", ""))
        except Exception:
            pass

    rating = (
        _safe_inner_text(page.locator('div.F7nice span[aria-hidden="true"]').first)
        or _first_attr(page, 'span[aria-label*="stars"]', "aria-label")
        or _safe_inner_text(page.locator('span[aria-label*="stars"]').first)
    )
    if rating:
        m = re.search(r"([0-9.]+)\s*stars", rating, re.I)
        rating = m.group(1) if m else _clean_whitespace(rating)

    reviews = (
        _first_attr_with_digits(page, 'button[aria-label*="reviews"]', "aria-label")
        or _extract_from_aria_label(page, "Reviews")
        or _extract_from_aria_label(page, "reviews")
    )
    if reviews:
        m = re.search(r"(\d[\d,]*)", reviews)
        reviews = m.group(1) if m else reviews

    address = _extract_from_aria_label(page, "Address") or _extract_from_data_item_id(page, "address")
    phone = _extract_from_aria_label(page, "Phone") or _extract_from_data_item_id(page, "phone")
    website = _extract_from_aria_label(page, "Website") or _extract_from_data_item_id(page, "authority")
    plus_code = _extract_from_aria_label(page, "Plus code") or _extract_from_data_item_id(page, "oloc")
    email = _extract_email_from_maps_page(page)

    place_lat, place_lng = _parse_coords_from_url(page.url)
    distance_km = None
    if place_lat is not None and place_lng is not None:
        distance_km = _haversine_km(parent_seed_lat, parent_seed_lng, place_lat, place_lng)

    return Lead(
        root_query=root_query,
        root_location=root_location,
        root_seed_name=root_seed_name,
        root_seed_lat=root_seed_lat,
        root_seed_lng=root_seed_lng,
        spawn_depth=depth,
        parent_seed_name=parent_seed_name,
        parent_seed_lat=parent_seed_lat,
        parent_seed_lng=parent_seed_lng,
        radius_km=radius_km,
        name=name,
        rating=rating,
        reviews=reviews,
        address=address,
        phone=phone,
        website=website,
        email=email,
        plus_code=plus_code,
        maps_url=page.url,
        place_lat=place_lat,
        place_lng=place_lng,
        distance_km_from_parent=distance_km,
    )


def write_xlsx(leads: Iterable[Lead], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "leads"

    leads_list = list(leads)
    if not leads_list:
        leads_list = [Lead(root_query="", root_location="")]

    headers = list(asdict(leads_list[0]).keys())
    ws.append(headers)
    for lead in leads_list:
        row = asdict(lead)
        ws.append([row.get(h) for h in headers])

    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=idx).value
            max_len = max(max_len, len(str(v or "")))
        ws.column_dimensions[col].width = min(max(12, max_len + 2), 70)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def spawn_scrape(
    *,
    root_query: str,
    root_location: str,
    nearby_query: str,
    radius_km: float,
    root_count: int,
    root_skip: int,
    per_seed_candidates: int,
    per_seed_keep_cap: int,
    max_depth: int,
    max_total: int,
    checkpoint_path: Optional[Path],
    checkpoint_every: int,
    headless: bool,
    slow_mo_ms: int,
    timeout_ms: int,
    zoom: int,
    json_stream: bool = False,
) -> list[Lead]:
    started = datetime.now(timezone.utc)
    LOG.info(
        'Spawn scrape: query="%s" location="%s" radius=%.2fkm max_total=%s max_depth=%s',
        root_query,
        root_location,
        radius_km,
        max_total,
        max_depth,
    )

    seed_search_text = f"{root_query} in {root_location}".strip()
    root_count = max(1, int(root_count))
    root_skip = max(0, int(root_skip))
    checkpoint_every = max(0, int(checkpoint_every))

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms)
        visited_urls: set[str] = set()
        kept: list[Lead] = []

        # First, collect the root URLs from the initial results list in a short-lived context.
        root_urls: list[str] = []
        try:
            ctx0 = browser.new_context(locale="en-GB", viewport={"width": 1400, "height": 900})
            page0 = ctx0.new_page()
            page0.set_default_timeout(timeout_ms)
            seed_url = f"https://www.google.com/maps/search/{quote_plus(seed_search_text)}"
            page0.goto(seed_url, wait_until="domcontentloaded", timeout=timeout_ms)
            _wait_for_results(page0, timeout_ms)
            root_urls = _collect_place_urls(page0, limit=root_skip + root_count)
            ctx0.close()
        except Exception:
            LOG.exception("Failed to collect root seeds from initial search")
            try:
                ctx0.close()
            except Exception:
                pass

        if not root_urls:
            raise RuntimeError("No place URLs found from the initial search results.")
        root_urls = root_urls[root_skip : root_skip + root_count]
        LOG.info("Collected %s root seed URL(s) from initial results (skip=%s)", len(root_urls), root_skip)

        # Collect multiple "root" places from the initial results list, then
        # spawn within radius around each until we hit max_total.
        # Seeds are place URLs with coords already parseable.
        Seed = tuple[str, str, float, float, int]  # (name, url, lat, lng, depth)

        def checkpoint_if_needed() -> None:
            if not checkpoint_path:
                return
            if checkpoint_every <= 0:
                return
            if len(kept) % checkpoint_every != 0:
                return
            try:
                write_xlsx(kept, checkpoint_path)
                LOG.info("Checkpoint: wrote %s lead(s) to %s", len(kept), checkpoint_path)
            except Exception:
                LOG.exception("Failed to write checkpoint workbook")

        for root_idx, root_url in enumerate(root_urls, start=1):
            if len(kept) >= max_total:
                break

            root_url = _normalise_maps_url(root_url)
            if root_url in visited_urls:
                continue

            # Use a fresh context per root; this dramatically reduces long-run Playwright flakiness.
            try:
                context = browser.new_context(locale="en-GB", viewport={"width": 1400, "height": 900})
                page = context.new_page()
                page.set_default_timeout(timeout_ms)
            except Exception:
                LOG.exception("Failed to create browser context")
                continue

            try:
                page.goto(root_url, wait_until="domcontentloaded", timeout=timeout_ms)
                page.wait_for_timeout(500)
            except Exception:
                LOG.exception("Failed to open root seed URL")
                try:
                    context.close()
                except Exception:
                    pass
                continue

            root_seed_name = (
                _safe_inner_text(page.locator("h1.DUwDvf").first)
                or _safe_inner_text(page.locator("h1").first)
                or "(unknown)"
            )
            root_lat, root_lng = _parse_coords_from_url(page.url)
            if root_lat is None or root_lng is None:
                LOG.info('Skipping root "%s" (no coords in URL)', root_seed_name)
                visited_urls.add(_normalise_maps_url(page.url))
                continue

            visited_urls.add(_normalise_maps_url(page.url))
            LOG.info('Root %s/%s: "%s" at (%.6f, %.6f)', root_idx, len(root_urls), root_seed_name, root_lat, root_lng)

            q: Deque[Seed] = deque()
            q.append((root_seed_name, page.url, root_lat, root_lng, 0))

            while q and len(kept) < max_total:
                parent_name, parent_url, parent_lat, parent_lng, depth = q.popleft()
                if depth >= max_depth:
                    continue

                elapsed_s = (datetime.now(timezone.utc) - started).total_seconds()
                LOG.info(
                    'Expanding root="%s" depth=%s seed="%s" (kept %s/%s, queue %s, elapsed %.1fs)',
                    root_seed_name,
                    depth,
                    parent_name,
                    len(kept),
                    max_total,
                    len(q),
                    elapsed_s,
                )

                nearby_url = f"https://www.google.com/maps/search/{quote_plus(nearby_query)}/@{parent_lat},{parent_lng},{zoom}z"
                page.goto(nearby_url, wait_until="domcontentloaded", timeout=timeout_ms)
                _wait_for_results(page, timeout_ms)

                candidate_urls = _collect_place_urls(page, limit=per_seed_candidates)

                within: list[tuple[str, float, float, float]] = []  # (url, dist_km, lat, lng)
                for url in candidate_urls:
                    url = _normalise_maps_url(url)
                    if url in visited_urls:
                        continue
                    lat, lng = _parse_coords_from_url(url)
                    if lat is None or lng is None:
                        continue
                    d = _haversine_km(parent_lat, parent_lng, lat, lng)
                    if d <= radius_km and d > 0.0:
                        within.append((url, d, lat, lng))

                within.sort(key=lambda t: t[1])

                scraped_this_seed = 0
                for url, dist, lat, lng in within:
                    if len(kept) >= max_total or scraped_this_seed >= per_seed_keep_cap:
                        break
                    if url in visited_urls:
                        continue

                    try:
                        page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
                        visited_urls.add(_normalise_maps_url(page.url))
                        lead = _scrape_place_page(
                            page,
                            root_query=root_query,
                            root_location=root_location,
                            root_seed_name=root_seed_name,
                            root_seed_lat=root_lat,
                            root_seed_lng=root_lng,
                            depth=depth + 1,
                            parent_seed_name=parent_name,
                            parent_seed_lat=parent_lat,
                            parent_seed_lng=parent_lng,
                            radius_km=radius_km,
                        )
                    except Exception:
                        LOG.exception("Failed to scrape place")
                        continue

                    if lead.distance_km_from_parent is None or lead.distance_km_from_parent > radius_km:
                        continue

                    kept.append(lead)
                    scraped_this_seed += 1
                    if json_stream:
                        print(json.dumps(asdict(lead)), flush=True)
                    checkpoint_if_needed()

                    if lead.place_lat is not None and lead.place_lng is not None and lead.name:
                        q.append(
                            (
                                lead.name,
                                _normalise_maps_url(lead.maps_url or url),
                                lead.place_lat,
                                lead.place_lng,
                                depth + 1,
                            )
                        )

            try:
                context.close()
            except Exception:
                pass

        browser.close()

        LOG.info("Finished. Kept %s lead(s). Visited %s URL(s).", len(kept), len(visited_urls))
        return kept


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Spawn/BFS scrape: expand within radius from discovered places.")
    parser.add_argument("--query", default="car wash", help='Initial seed query, e.g. "car wash"')
    parser.add_argument("--location", default="Delaware", help='Initial seed location, e.g. "Delaware"')
    parser.add_argument("--nearby-query", default=None, help="Nearby search query (defaults to --query)")
    parser.add_argument("--radius-km", type=float, default=1.0, help="Radius in kilometres (default 1.0)")
    parser.add_argument("--max-total", type=int, default=50, help="Max total leads to keep (default 50)")
    parser.add_argument("--root-count", type=int, default=50, help="How many initial search results to use as root seeds (default 50)")
    parser.add_argument("--root-skip", type=int, default=0, help="How many initial search results to skip before taking roots (default 0)")
    parser.add_argument("--max-depth", type=int, default=3, help="Max spawn depth (default 3)")
    parser.add_argument("--per-seed-candidates", type=int, default=60, help="Candidate URLs to collect per seed (default 60)")
    parser.add_argument("--per-seed-keep-cap", type=int, default=10, help="Max leads to keep per seed expansion (default 10)")
    parser.add_argument("--zoom", type=int, default=15, help="Zoom used for nearby search (default 15)")
    parser.add_argument("--checkpoint-every", type=int, default=10, help="Write partial results every N leads (default 10, 0 disables)")
    parser.add_argument("--headless", action="store_true", help="Run headless (default is headed)")
    parser.add_argument("--slowmo-ms", type=int, default=50, help="Slow motion delay per action (ms)")
    parser.add_argument("--timeout-ms", type=int, default=90_000, help="Playwright timeout (ms)")
    parser.add_argument("--json-stream", action="store_true", help="Emit each lead as a JSON line to stdout (for SSE integration)")
    parser.add_argument("--out", default="spawn-radius-leads.xlsx", help="Output .xlsx path")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    nearby_query = args.nearby_query if args.nearby_query is not None else args.query
    out_path = Path(args.out).expanduser().resolve()

    leads = spawn_scrape(
        root_query=args.query,
        root_location=args.location,
        nearby_query=nearby_query,
        radius_km=args.radius_km,
        root_count=args.root_count,
        root_skip=args.root_skip,
        per_seed_candidates=args.per_seed_candidates,
        per_seed_keep_cap=args.per_seed_keep_cap,
        max_depth=args.max_depth,
        max_total=args.max_total,
        checkpoint_path=out_path,
        checkpoint_every=args.checkpoint_every,
        headless=args.headless,
        slow_mo_ms=args.slowmo_ms,
        timeout_ms=args.timeout_ms,
        zoom=args.zoom,
        json_stream=args.json_stream,
    )

    write_xlsx(leads, out_path)
    print(f"Wrote {len(leads)} lead(s) to: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

